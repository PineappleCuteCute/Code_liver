import openpyxl
import numpy as np
import time


def importData(sheet):
    return ([[cell.value for cell in row] for row in sheet])


wb = openpyxl.load_workbook("Data.xlsx")


def importFuzzyData(data):
    fuzzyData = wb[data]
    base = list(fuzzyData.values)

    return base


def combination(k, n):
    if k == 0 or k == n:
        return 1
    if k == 1:
        return n
    return combination(k - 1, n - 1) + combination(k, n - 1)


def caculateA(base):
    colum = len(base[0])
    row = len(base)
    A = np.zeros((row, combination(4, colum - 1)))

    for r1 in range(row):
        k = [0] * combination(4, colum - 1)
        temp = 0
        for a in range(0, colum - 4):
            for b in range(a + 1, colum - 3):
                for c in range(b + 1, colum - 2):
                    for d in range(c + 1, colum - 1):
                        for r2 in range(row):
                            if base[r1][a] == base[r2][a] and base[r1][b] == base[r2][b] and base[r1][c] == base[r2][
                                c] and base[r1][d] == base[r2][d]:
                                k[temp] += 1

                        A[r1][temp] = k[temp] / row
                        temp += 1
    print("done A")
    return A


"""A = (caculateA(base))

for i in range(len(A)):
    print(sum(A[i]))"""


def caculateM(base):
    colum = len(base[0])
    row = len(base)
    M = np.zeros((row, colum - 1))
    for t1 in range(row):
        k = [0] * (colum - 1)
        temp = 0
        for i in range(colum - 1):
            for t2 in range(row):
                if base[t1][i] == base[t2][i] and base[t1][colum - 1] == base[t2][colum - 1]:
                    k[temp] += 1
            M[t1][temp] = k[temp] / row
            temp += 1

    return M


# print(caculateM(base))


def caculateB(base, A, M):
    colum = len(base[0])
    row = len(base)
    B = np.zeros((row, combination(3, colum - 1)))

    for r in range(row):
        temp = 0
        for a in range(0, colum - 3):
            for b in range(a + 1, colum - 2):
                for c in range(b + 1, colum - 1):
                    B[r][temp] = sum(A[r]) * min(M[r][a], M[r][b], M[r][c])
                    temp += 1
    print("done B")
    return B


# print(caculateB(base, caculateA(base), caculateM(base)))
# print(caculateM(base))


def caculateC(base, B):
    colum = len(base[0])
    row = len(base)
    cols = 2 * combination(3, colum - 1)
    C = np.zeros((row, cols))

    for r1 in range(row):
        temp = 0
        for i in range(2):
            for a in range(0, (colum - 3)):
                for b in range(a + 1, (colum - 2)):
                    for c in range(b + 1, (colum - 1)):
                        for r2 in range(row):
                            if base[r1][a] == base[r2][a] and base[r1][b] == base[r2][b] and base[r1][c] == base[r2][
                                c] and base[r2][colum - 1] == i:
                                C[r1][temp] += B[r2][temp % combination(3, colum - 1)]
                        # print(temp,":",temp//combination(3,colum-1))
                        temp += 1
    print("done C")
    return C


# print(caculateC(base, caculateB(base, caculateA(base), caculateM(base)))[0])

def writeToExcel(sheet, table, row):
    for x in range(row):
        for y in range(len(table[x])):
            sheet.cell(row=x + 1, column=y + 1, value=table[x][y])


def update(data):
    # wb = openpyxl.load_workbook("FKG_Py.xlsx")
    base = importFuzzyData(data)
    startA = time.time()
    A = caculateA(base)
    print("Time A: ", time.time() - startA)
    startB = time.time()
    M = caculateM(base)
    B = caculateB(base, A, M)
    print("Time B: ", time.time() - startB)
    startC = time.time()
    C = caculateC(base, B)
    print("Time C: ", time.time() - startC)
    # writeToExcel(wb['A'], A, len(A))
    # writeToExcel(wb['M'], M, len(M))
    # writeToExcel(wb['B'], B, len(B))
    writeToExcel(wb['C'], C, len(C))

    wb.save("Data.xlsx")
    print("Update successful! Let's Start")


def FISA(base, C, list):
    colum = len(base[0])
    row = len(base)

    cols = combination(3, (colum - 1))
    C0 = [0] * cols
    C1 = [0] * cols

    t = 0
    for a in range(0, colum - 3):
        for b in range(a + 1, colum - 2):
            for c in range(b + 1, colum - 1):
                for r in range(row - 1):
                    if base[r][a] == list[a] and base[r][b] == list[b] and base[r][c] == list[c] and base[r][
                        colum - 1] == 0:
                        C0[t] = C[r][t + 0 * cols]
                        # break
                    if base[r][a] == list[a] and base[r][b] == list[b] and base[r][c] == list[c] and base[r][
                        colum - 1] == 1:
                        C1[t] = C[r][t + 1 * cols]
                        # break
                t += 1
    # print(t)

    D0 = max(C0) + min(C0)
    D1 = max(C1) + min(C1)

    # print(D0, max(C0), min(C0))
    # print(D1, max(C1), min(C1))
    # print(D2, max(C2), min(C2))
    if D0 > D1:
        return 0
    else:
        return 1


def Acc(A, B):
    result = 0

    for i in range(len(A)):
        if int(A[i]) - int(B[i]) == 0:
            result += 1

    return round(result * 100 / len(A), 2)


def Tprecision(Pre, Act):
    result = 0
    TP = 0
    FP = 0

    for i in range(len(Pre)):
        if int(Pre[i]) == 1 and int(Act[i]) == 1:
            TP += 1
        if int(Pre[i]) == 1 and int(Act[i]) == 0:
            FP += 1

    # return str(TP) +  " : " + str(TP+FP)

    if TP:
        return round(100 * TP / (TP + FP), 2)
    else:
        if FP:
            return 0
        else:
            return None


def Trecall(Pre, Act):
    result = 0
    TP = 0
    FN = 0

    for i in range(len(Pre)):
        if int(Pre[i]) == 1 and int(Act[i]) == 1:
            TP += 1
        if int(Pre[i]) == 0 and int(Act[i]) == 1:
            FN += 1

    # return str(TP) +  " : " + str(TP+FN)

    if TP:
        return round(100 * TP / (TP + FN), 2)
    else:
        if FN:
            return 0
        else:
            return None


listAcc = []
listPre = []
listRe = []
timeTest = []
timeUpdate = []


def testAccuracy(data, Te):
    test = importData(wb[Te])
    sheetC = wb['C']
    base = importFuzzyData(data)
    # print(base)
    C = list(sheetC.values)
    X = np.zeros(len(test))
    X_test = np.array(test).T[-1]
    for i in range(len(test)):
        X[i] = FISA(base, C, test[i])
        # print(test[i])

    # print(X)
    # print(X_test)
    listAcc.append(Acc(X, X_test))
    listPre.append(Tprecision(X, X_test))
    listRe.append(Trecall(X, X_test))
    # wb.save("Data.xlsx")


for i in range(1, 21):
    data = "D0".replace('0', str(i))
    test = "T0".replace('0', str(i))
    print(data, ':', test)
    start = time.time()
    update(data)
    timeUpdate.append(round(time.time() - start, 2))
    # print("Time update: ", time.time() - start)
    start2 = time.time()
    testAccuracy(data, test)
    timeTest.append(round(time.time() - start2, 2))
    # print("Time test: ", time.time() - start2)

print(timeUpdate)
print(timeTest)
print(listAcc)

RESULTS = []
RESULTS.append(timeUpdate)
RESULTS.append(timeTest)
RESULTS.append(listAcc)
RESULTS.append(listPre)
RESULTS.append(listRe)
writeToExcel(wb['results'], RESULTS, len(RESULTS))
wb.save("Data.xlsx")


